"""
Export des resultats au format Excel compatible Odoo 18.

Ce module genere un fichier .xlsx avec deux onglets :
- "product.product" : tableau des produits avec les images associees
- "Resume"          : rapport de synthese (date, stats, formules)

Le fichier est directement importable dans Odoo 18.
"""

from datetime import datetime  # Pour horodater le rapport

import openpyxl                                 # Manipulation de fichiers Excel
from openpyxl.styles import Font, PatternFill, Alignment  # Styles des cellules
from openpyxl.utils import get_column_letter    # Convertit un numero de colonne en lettre (1 -> A)

# Import des noms de colonnes Odoo depuis la config
from .config import ODOO_SKU_COLUMN, ODOO_IMG_COLUMN, ODOO_EXT_COLUMN, ODOO_VAR_COLUMN
from .logger import log  # Logging thread-safe


def export_odoo_excel(df, sku_files, output_path):
    """
    Genere un fichier Excel formate pour l'import Odoo 18.

    Le fichier contient :
    - Un onglet "product.product" avec les donnees produit et les noms de fichiers images
    - Un onglet "Resume" avec les statistiques du scraping (total, succes, echecs)

    Les cellules sont colorees selon l'etat :
    - Vert  : image trouvee
    - Jaune : image manquante
    - Gris  : lignes alternees pour la lisibilite

    Args:
        df:          DataFrame pandas contenant les donnees produit du fichier source.
        sku_files:   Dict {sku: [filename, ...]} associant chaque SKU a ses fichiers images.
        output_path: Chemin Path du fichier Excel de sortie.
    """
    # Copie le DataFrame pour ne pas modifier l'original
    out = df.copy()
    # Si la colonne image n'existe pas dans le DataFrame, on la cree
    if ODOO_IMG_COLUMN not in out.columns:
        out[ODOO_IMG_COLUMN] = ""

    # Remplit la colonne image avec le premier fichier trouve pour chaque SKU
    # Lambda : recupere la liste des fichiers pour le SKU, prend le premier (ou "")
    out[ODOO_IMG_COLUMN] = out[ODOO_SKU_COLUMN].apply(
        lambda sku: (sku_files.get(str(sku).strip()) or [""])[0]
    )

    # ---- Creation du classeur Excel ----
    # Cree un nouveau classeur vide
    wb = openpyxl.Workbook()
    # Recupere la feuille active (premiere feuille creee automatiquement)
    ws = wb.active
    # Renomme la feuille en "product.product" (nom attendu par Odoo)
    ws.title = "product.product"

    # ---- Style de l'en-tete ----
    # Liste des noms de colonnes du DataFrame
    headers = list(out.columns)
    # Fond bleu fonce pour l'en-tete
    hdr_fill = PatternFill("solid", start_color="1D3557")
    # Texte blanc gras en Arial 10 pour l'en-tete
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    # Alignement centre horizontal et vertical avec retour a la ligne
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ecrit les en-tetes dans la premiere ligne
    for ci, col in enumerate(headers, 1):
        # Ecrit le nom de la colonne dans la cellule (ligne 1, colonne ci)
        c = ws.cell(row=1, column=ci, value=col)
        # Applique le fond bleu
        c.fill = hdr_fill
        # Applique le texte blanc gras
        c.font = hdr_font
        # Applique l'alignement centre
        c.alignment = hdr_align

    # ---- Styles des donnees ----
    # Fond gris clair pour les lignes paires (alternance zebra)
    alt_fill = PatternFill("solid", start_color="F0F4F8")
    # Fond vert clair pour les cellules avec image trouvee
    ok_fill = PatternFill("solid", start_color="D4EDDA")
    # Fond jaune clair pour les cellules avec image manquante
    miss_fill = PatternFill("solid", start_color="FFF3CD")
    # Police Arial 9 pour toutes les cellules de donnees
    norm_font = Font(name="Arial", size=9)
    # Index de la colonne image (1-indexed pour openpyxl)
    img_ci = headers.index(ODOO_IMG_COLUMN) + 1

    # ---- Ecriture des donnees ligne par ligne ----
    # Parcourt chaque ligne du DataFrame (ri commence a 2 car ligne 1 = en-tete)
    for ri, row in enumerate(out.itertuples(index=False), 2):
        # Parcourt chaque valeur de la ligne
        for ci, value in enumerate(row, 1):
            # Convertit NaN et None en chaine vide pour l'affichage
            val = "" if str(value) in ("nan", "None") else str(value)
            # Ecrit la valeur dans la cellule
            c = ws.cell(row=ri, column=ci, value=val)
            # Applique la police standard
            c.font = norm_font
            # Alignement vertical centre
            c.alignment = Alignment(vertical="center")
            # Coloration conditionnelle de la colonne image
            if ci == img_ci:
                # Vert si image presente, jaune si manquante
                c.fill = ok_fill if val else miss_fill
            elif ri % 2 == 0:
                # Gris clair pour les lignes paires (effet zebra)
                c.fill = alt_fill

    # ---- Largeurs des colonnes ----
    # Largeurs personnalisees pour les colonnes Odoo
    col_widths = {
        ODOO_EXT_COLUMN: 45,  # External ID : large pour les identifiants longs
        ODOO_SKU_COLUMN: 20,  # Internal Reference : taille moyenne
        ODOO_VAR_COLUMN: 18,  # Variant Values : taille moyenne
        ODOO_IMG_COLUMN: 30,  # Variant Image : large pour les noms de fichiers
    }
    # Applique les largeurs (20 par defaut si la colonne n'est pas dans le dict)
    for ci, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 20)

    # Hauteur de la ligne d'en-tete (28 pixels)
    ws.row_dimensions[1].height = 28
    # Fige la premiere ligne (l'en-tete reste visible au scroll)
    ws.freeze_panes = "A2"

    # ---- Onglet Resume (statistiques) ----
    # Lettre de la colonne image (pour les formules COUNTIF)
    img_col_letter = get_column_letter(img_ci)
    # Nombre total de lignes de donnees (pour la plage des formules)
    n_rows = len(out) + 1
    # Cree un deuxieme onglet pour le rapport de synthese
    ws2 = wb.create_sheet("Resume")
    # Titre du rapport
    ws2["A1"] = "Rapport de Scraping"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial")
    # Date et heure de generation du rapport
    ws2["A3"] = "Date de generation"
    ws2["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    # Nombre total de SKUs traites
    ws2["A4"] = "Total SKUs"
    ws2["B4"] = len(sku_files)
    # Nombre de SKUs avec au moins une image trouvee
    ws2["A5"] = "Images trouvees"
    ws2["B5"] = sum(1 for v in sku_files.values() if v)
    # Nombre de SKUs sans image (formule Excel : total - trouvees)
    ws2["A6"] = "Images manquantes"
    ws2["B6"] = "=B4-B5"
    # Nombre de variantes avec un fichier image valide (formule COUNTIF)
    ws2["A7"] = "Total Variantes OK"
    ws2["B7"] = (
        # Compte les cellules contenant .jpg dans la colonne image
        f"=COUNTIF('product.product'!{img_col_letter}2:{img_col_letter}{n_rows},\"*.jpg\")"
        # + celles contenant .png
        f"+COUNTIF('product.product'!{img_col_letter}2:{img_col_letter}{n_rows},\"*.png\")"
        # + celles contenant .webp
        f"+COUNTIF('product.product'!{img_col_letter}2:{img_col_letter}{n_rows},\"*.webp\")"
    )
    # Applique le style gras aux labels du rapport
    for cell in ["A3", "A4", "A5", "A6", "A7"]:
        ws2[cell].font = Font(bold=True, name="Arial", size=10)
    # Largeur de la colonne A (labels)
    ws2.column_dimensions["A"].width = 22
    # Largeur de la colonne B (valeurs)
    ws2.column_dimensions["B"].width = 28

    # Sauvegarde le classeur sur le disque
    wb.save(str(output_path))
    # Log du chemin absolu du fichier genere
    log(f"Fichier Odoo sauvegarde : {output_path.resolve()}", "OK")
