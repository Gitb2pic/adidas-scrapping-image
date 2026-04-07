#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║      Image Scraper Générique — 3 modes                       ║
╠══════════════════════════════════════════════════════════════╣
║  MODE 1  SKU unique                                          ║
║    python scraper.py --sku MH03AC-T03RED                     ║
║                                                              ║
║  MODE 2  Batch depuis Excel  +  export Odoo 18               ║
║    python scraper.py --excel variants.xlsx                   ║
║                                                              ║
║  MODE 3  Multi-URL async (course, premier gagne)             ║
║    python scraper.py --sku MH03AC-T03RED --urls              ║
║        https://mostlyheardrarelyseen8bit.com/search?q={sku}  ║
║                                                              ║
║  Options : --output ./images  --headless  --dry-run          ║
╚══════════════════════════════════════════════════════════════╝

Requirements:
    pip install selenium webdriver-manager openpyxl pandas
"""

import re
import sys
import json
import time
import argparse
import threading
import urllib.request
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from webdriver_manager.chrome import ChromeDriverManager
except ImportError:
    print("\n[X] Dependances manquantes. Lancer d'abord :\n")
    print("    pip install selenium webdriver-manager openpyxl pandas\n")
    sys.exit(1)

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


# ================================================================
# CONFIG
# ================================================================

DEFAULT_URLS = [
    "https://www.google.com/search?q={sku}+product&udm=2",
    "https://mostlyheardrarelyseen8bit.com/search?q={sku}",
    "https://www.farfetch.com/be/search?q={sku}",
]

DOWNLOAD_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.google.com/",
}

ODOO_SKU_COLUMN = "Internal Reference"
ODOO_IMG_COLUMN = "Variant Image"
ODOO_EXT_COLUMN = "External ID"
ODOO_VAR_COLUMN = "Variant Values"


# ================================================================
# LOGGING (thread-safe)
# ================================================================

_log_lock = threading.Lock()

def log(msg, level="INFO", sku=""):
    icons = {"INFO": "i ", "OK": "OK", "WARN": "! ", "ERR": "X ", "DL": "->"}
    prefix = f"[{sku}] " if sku else ""
    with _log_lock:
        print(f"  [{icons.get(level, '  ')}] {prefix}{msg}")

def banner(text):
    bar = "=" * 60
    print(f"\n{bar}\n  {text}\n{bar}\n")


# ================================================================
# VALIDATION ET FORMATAGE D'IMAGE
# ================================================================

def get_high_res_url(url: str) -> str:
    """
    Supprime les suffixes de taille typiques de Shopify pour forcer 
    le telechargement de l'image source en haute resolution.
    """
    if not url: return ""
    return re.sub(r'_([0-9]+x[0-9]*|[0-9]*x[0-9]+)(\.[a-zA-Z0-9]{3,4})', r'\2', url, flags=re.IGNORECASE)

def is_valid_image(url: str) -> bool:
    """
    Valide qu'une URL est bien une image de produit et non une icone.
    """
    url_lower = url.lower()
    clean = url.split("?")[0]

    excluded_patterns = [
        "/logo", "/icon", "/badge", "/banner",
        "/flag", "/avatar", "/favicon", "/sprite",
        "placeholder", "fallback", "default",
    ]
    if any(p in url_lower for p in excluded_patterns):
        return False

    if re.search(r'\.(jpg|jpeg|png|webp)$', clean, re.I):
        return True
        
    if "cdn.shopify.com" in url_lower or "/products/" in url_lower or "/files/" in url_lower:
        return True

    return False


# ================================================================
# SELENIUM DRIVER
# ================================================================

def build_driver(headless):
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--window-size=1400,900")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"}
    )
    return driver


def accept_cookies(driver):
    try:
        btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR,
                "button#onetrust-accept-btn-handler, "
                "button[data-testid='cookie-accept-button'], "
                "button.btn-accept-all"
            ))
        )
        btn.click()
        time.sleep(1)
    except Exception:
        pass


# ================================================================
# EXTRACTION D'IMAGES
# ================================================================

def extract_images(driver, sku):
    """
    Extrait toutes les images. Priorité STRICTE au HTML (de haut en bas) 
    pour garantir que l'image [0] est bien la toute premiere vue sur la page.
    """
    images = []
    seen   = set()

    def add(url, label):
        if url.startswith("//"):
            url = "https:" + url
            
        url = get_high_res_url(url)
        clean = url.split("?")[0]
        
        if clean in seen:
            return
        seen.add(clean)
        
        if is_valid_image(clean):
            images.append({"url": url, "label": label})

    # 1. HTML DOM (Priorité absolue : premier element lu = première image retenue)
    try:
        # On cible d'abord les conteneurs de produits évidents
        selectors = [
            "img[class*='product']",
            "img[class*='grid']",
            "img[class*='card']",
            "a[class*='product'] img",
            "img" # Capture générale en dernier recours
        ]
        
        for sel in selectors:
            for el in driver.find_elements(By.CSS_SELECTOR, sel):
                for attr in ("src", "data-src", "srcset", "data-srcset"):
                    val = el.get_attribute(attr) or ""
                    if "srcset" in attr and val:
                        for part in val.split(","):
                            part_url = part.strip().split(" ")[0]
                            if part_url.startswith("http") or part_url.startswith("//"):
                                add(part_url, "html_srcset")
                    else:
                        if val.startswith("http") or val.startswith("//"):
                            add(val, "html_img")
                            
        if images:
            log(f"{len(images)} image(s) valide(s) extraite(s) du visuel HTML", "OK", sku=sku)
            return images
    except Exception:
        pass

    # 2. Objets JSON imbriques (Plan B si le HTML échoue)
    for var in ("__NEXT_DATA__", "__STATE__", "__PRELOADED_STATE__", "__INITIAL_STATE__"):
        try:
            if var == "__NEXT_DATA__":
                el = driver.find_element(By.ID, var)
                raw = el.get_attribute("textContent")
            else:
                raw = driver.execute_script(f"return JSON.stringify(window.{var})")
            
            if raw:
                _walk_json(json.loads(raw), add)
                if images:
                    log(f"{len(images)} image(s) extraite(s) des donnees cachees {var}", "OK", sku=sku)
                    return images
        except Exception:
            pass

    # 3. Regex de secours sur le code source brut
    cdn_re = re.compile(
        r'(?:https?:)?//[^\s"\'\\<>]+\.(?:jpg|jpeg|png|webp)',
        re.I
    )
    for m in cdn_re.finditer(driver.page_source):
        add(m.group(0), "regex_cdn")

    if images:
        log(f"{len(images)} image(s) valide(s) recuperee(s) par Regex", "OK", sku=sku)
    else:
        log("Aucune image valide trouvee sur cette page.", "WARN", sku=sku)

    return images


def _walk_json(node, callback):
    if isinstance(node, dict):
        for key in ("view_list", "images", "gallery_images", "media", "src", "url"):
            val = node.get(key)
            if isinstance(val, list):
                for i, item in enumerate(val):
                    if isinstance(item, str) and (item.startswith("http") or item.startswith("//")):
                        callback(item, f"{key}_{i}")
                    elif isinstance(item, dict):
                        for k in ("image_url", "src", "url", "href"):
                            u = item.get(k, "")
                            if u and (u.startswith("http") or u.startswith("//")):
                                callback(u, f"{key}_{i}_{k}")
                                break
            elif isinstance(val, str) and (val.startswith("http") or val.startswith("//")):
                callback(val, f"json_direct_{key}")
                
        for v in node.values():
            _walk_json(v, callback)
    elif isinstance(node, list):
        for item in node:
            _walk_json(item, callback)


# ================================================================
# MODE 1 : Navigation sequentielle (tout site)
# ================================================================

# Selecteurs generiques pour detecter un lien produit sur n'importe quel site
PRODUCT_LINK_SELECTORS = [
    # Selecteurs specifiques e-commerce
    "a.grid-view-item__link",
    "a.product-card",
    "a.grid-product__link",
    "a.product-item__link",
    "a[data-testid='product-card-link']",
    "a[class*='product-card']",
    "div[class*='product-card'] a",
    "a[class*='product-link']",
    "a[class*='product-item']",
    "a[class*='product-tile']",
    # Selecteurs generiques (tout site)
    "a[href*='/product']",
    "a[href*='/products/']",
    "a[href*='/p/']",
    "a[href*='/item/']",
    "a[href*='/dp/']",
    "[class*='product'] a[href]",
    "[class*='item'] a[href]",
    "[class*='card'] a[href]",
    "[class*='tile'] a[href]",
    "[data-product] a[href]",
    "[data-item] a[href]",
]


def navigate_single(driver, sku, url=None):
    search = (url or DEFAULT_URLS[0]).replace("{sku}", sku.upper())
    log(f"Recherche : {search}", sku=sku)
    driver.get(search)
    accept_cookies(driver)
    time.sleep(3)

    # Essayer de cliquer sur le premier produit trouve
    for selector in PRODUCT_LINK_SELECTORS:
        try:
            el = driver.find_element(By.CSS_SELECTOR, selector)
            href = el.get_attribute("href")
            if href and href.startswith("http"):
                log(f"Produit detecte : {href}", sku=sku)
                driver.get(href)
                time.sleep(3)
                return True
        except Exception:
            continue

    log("Extraction directe depuis la page de recherche.", "INFO", sku=sku)
    return True


# ================================================================
# MODE 3 : Multi-URL ASYNC (course, premier gagnant)
# ================================================================

class RaceResult:
    def __init__(self):
        self.winner_url    = None
        self.winner_images = []
        self.lock          = threading.Lock()
        self.found         = threading.Event()

    def claim(self, url, images):
        with self.lock:
            if not self.found.is_set():
                self.winner_url    = url
                self.winner_images = images
                self.found.set()
                return True
        return False


def _race_worker(url, sku, headless, race):
    if race.found.is_set():
        return

    driver = None
    try:
        driver = build_driver(headless)
        log(f"[RACE] -> {url}", sku=sku)
        driver.get(url)
        accept_cookies(driver)
        time.sleep(3)

        if "404" in driver.title.lower():
            log(f"[RACE] Erreur 404 : {url}", "WARN", sku=sku)
            return

        if race.found.is_set():
            return

        # Essayer de naviguer vers une page produit
        for selector in PRODUCT_LINK_SELECTORS:
            try:
                el = driver.find_element(By.CSS_SELECTOR, selector)
                href = el.get_attribute("href")
                if href and href.startswith("http"):
                    log(f"[RACE] Produit detecte : {href}", sku=sku)
                    driver.get(href)
                    time.sleep(3)
                    break
            except Exception:
                continue

        if race.found.is_set():
            return

        images = extract_images(driver, sku)

        if images and not race.found.is_set():
            if race.claim(url, images):
                log(f"[RACE] SUCCES sur : {url}  ({len(images)} images)", "OK", sku=sku)
        else:
            log(f"[RACE] Aucune image : {url}", "WARN", sku=sku)

    except Exception as e:
        if not race.found.is_set():
            log(f"[RACE] Erreur {url} : {e}", "WARN", sku=sku)
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


def navigate_race(sku, urls, headless, timeout=90):
    expanded = [u.replace("{sku}", sku.upper()) for u in urls]
    race = RaceResult()
    log(f"[RACE] {len(expanded)} URLs en parallele", sku=sku)

    with ThreadPoolExecutor(max_workers=len(expanded)) as pool:
        for url in expanded:
            pool.submit(_race_worker, url, sku, headless, race)
        race.found.wait(timeout=timeout)

    if not race.found.is_set():
        log("Echec (timeout ou aucune source valide).", "ERR", sku=sku)
        return []

    return race.winner_images


# ================================================================
# TELECHARGEMENT
# ================================================================

def download_image(url, dest):
    req = urllib.request.Request(url, headers=DOWNLOAD_HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            dest.write_bytes(resp.read())
        return True
    except Exception as e:
        log(f"Echec du telechargement {dest.name} : {e}", "ERR")
        return False


def download_all(images, output_dir, sku):
    """
    Telecharge TOUTES les images du produit.
    Retourne un dict avec :
      - "all": liste de tous les fichiers telecharges
      - "preferred": le fichier prefere pour Excel (images[1] sinon images[0])
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    downloaded = []
    preferred  = None

    for idx, img in enumerate(images):
        url = img["url"]
        clean_url = url.split("?")[0]
        ext = clean_url.rsplit(".", 1)[-1] if "." in clean_url else "jpg"
        ext = ext if ext.lower() in {"jpg", "jpeg", "png", "webp"} else "jpg"

        fname = f"{sku.upper()}_{idx}.{ext}"
        dest  = output_dir / fname

        log(f"Telechargement image[{idx}]: {fname}", "DL", sku=sku)
        if download_image(url, dest):
            log(f"OK : {dest.stat().st_size // 1024} KB", "OK", sku=sku)
            entry = {"filename": fname, "path": dest, "url": url, "index": idx}
            downloaded.append(entry)
            # Priorite a images[1] pour Excel, fallback images[0]
            if idx == 1:
                preferred = entry
            elif idx == 0 and preferred is None:
                preferred = entry
        else:
            log(f"Echec image[{idx}]", "WARN", sku=sku)

    if downloaded:
        log(f"{len(downloaded)}/{len(images)} image(s) telechargee(s)", "OK", sku=sku)
    else:
        log("Impossible de sauvegarder aucun fichier.", "ERR", sku=sku)

    return {"all": downloaded, "preferred": preferred}


# ================================================================
# EXPORT ODOO 18
# ================================================================

def export_odoo_excel(df, sku_files, output_path):
    out = df.copy()
    if ODOO_IMG_COLUMN not in out.columns:
        out[ODOO_IMG_COLUMN] = ""

    out[ODOO_IMG_COLUMN] = out[ODOO_SKU_COLUMN].apply(
        lambda sku: (sku_files.get(str(sku).strip()) or [""])[0]
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "product.product"

    headers   = list(out.columns)
    hdr_fill  = PatternFill("solid", start_color="1D3557")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, col in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = hdr_align

    alt_fill  = PatternFill("solid", start_color="F0F4F8")
    ok_fill   = PatternFill("solid", start_color="D4EDDA")
    miss_fill = PatternFill("solid", start_color="FFF3CD")
    norm_font = Font(name="Arial", size=9)
    img_ci    = headers.index(ODOO_IMG_COLUMN) + 1

    for ri, row in enumerate(out.itertuples(index=False), 2):
        for ci, value in enumerate(row, 1):
            val = "" if str(value) in ("nan", "None") else str(value)
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = norm_font
            c.alignment = Alignment(vertical="center")
            if ci == img_ci:
                c.fill = ok_fill if val else miss_fill
            elif ri % 2 == 0:
                c.fill = alt_fill

    col_widths = {
        ODOO_EXT_COLUMN: 45,
        ODOO_SKU_COLUMN: 20,
        ODOO_VAR_COLUMN: 18,
        ODOO_IMG_COLUMN: 30,
    }
    for ci, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 20)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    img_col_letter = get_column_letter(img_ci)
    n_rows = len(out) + 1
    ws2 = wb.create_sheet("Resume")
    ws2["A1"] = "Rapport de Scraping"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial")
    ws2["A3"] = "Date de generation"; ws2["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws2["A4"] = "Total SKUs"        ; ws2["B4"] = len(sku_files)
    ws2["A5"] = "Images trouvees"   ; ws2["B5"] = sum(1 for v in sku_files.values() if v)
    ws2["A6"] = "Images manquantes" ; ws2["B6"] = f"=B4-B5"
    ws2["A7"] = "Total Variantes OK"
    ws2["B7"] = (
        f"=COUNTIF('product.product'!{img_col_letter}2:{img_col_letter}{n_rows},\"*.jpg\")"
        f"+COUNTIF('product.product'!{img_col_letter}2:{img_col_letter}{n_rows},\"*.png\")"
        f"+COUNTIF('product.product'!{img_col_letter}2:{img_col_letter}{n_rows},\"*.webp\")"
    )
    for cell in ["A3","A4","A5","A6","A7"]:
        ws2[cell].font = Font(bold=True, name="Arial", size=10)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 28

    wb.save(str(output_path))
    log(f"Fichier Odoo sauvegarde : {output_path.resolve()}", "OK")


# ================================================================
# ORCHESTRATION
# ================================================================

def scrape_sku(sku, output_base, headless, dry_run, use_race, custom_urls):
    urls = custom_urls if custom_urls else DEFAULT_URLS

    if use_race or len(urls) > 1:
        images = navigate_race(sku, urls, headless)
    else:
        driver = build_driver(headless)
        try:
            if not navigate_single(driver, sku, url=urls[0] if urls else None):
                return {"all": [], "preferred": None}
            accept_cookies(driver)
            time.sleep(2)
            images = extract_images(driver, sku)
        finally:
            driver.quit()

    if not images:
        log("Processus termine : Aucune image", "WARN", sku=sku)
        return {"all": [], "preferred": None}

    if dry_run:
        log(f"[DRY-RUN] Cible detectee : {images[0]['url']}", "OK", sku=sku)
        return {"all": [{"filename": "", "url": img["url"]} for img in images],
                "preferred": {"filename": "", "url": images[1]["url"] if len(images) > 1 else images[0]["url"]}}

    return download_all(images, output_base, sku)


def run_single(sku, output_base, headless, dry_run, use_race, custom_urls):
    mode = "RACE (Analyse multi-URL)" if use_race else "Sequentiel"
    banner(f"MODE 1 — SKU : {sku.upper()}  [{mode}]")
    result = scrape_sku(sku, output_base, headless, dry_run, use_race, custom_urls)
    if not result or not result.get("all"):
        sys.exit(1)


def run_excel_batch(excel_path, output_base, headless, dry_run, use_race, custom_urls):
    if not HAS_PANDAS:
        print("[X] Des bibliotheques sont manquantes. Lancer : pip install pandas openpyxl")
        sys.exit(1)
    banner("MODE 2 — Lecture Excel et creation du fichier Odoo")

    df = pd.read_excel(excel_path, dtype=str)
    if ODOO_SKU_COLUMN not in df.columns:
        print(f"[X] Introuvable : Colonne '{ODOO_SKU_COLUMN}'. Options detectees : {list(df.columns)}")
        sys.exit(1)

    skus = df[ODOO_SKU_COLUMN].dropna().unique().tolist()
    log(f"Demarrage sur {len(skus)} element(s) unique(s) : {skus}")

    sku_files = {}

    for i, sku in enumerate(skus, 1):
        sku = str(sku).strip()
        print(f"\n[{i}/{len(skus)}] Element: {sku} {'─'*35}")
        result = scrape_sku(sku, output_base, headless, dry_run, use_race, custom_urls)
        # Pour Excel : uniquement l'image preferee (images[1] ou fallback images[0])
        pref = result.get("preferred") if result else None
        sku_files[sku] = [pref["filename"]] if pref and pref.get("filename") else []

    if not dry_run:
        output_base.mkdir(parents=True, exist_ok=True)
        ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
        odoo_path = output_base / f"odoo18_import_{ts}.xlsx"
        export_odoo_excel(df, sku_files, odoo_path)

        found  = sum(1 for v in sku_files.values() if v)
        missed = [s for s, v in sku_files.items() if not v]
        print(f"\n  Bilan du traitement : {found}/{len(skus)} succes.")
        if missed:
            print(f"  Elements non resolus : {', '.join(missed)}")
        print(f"  Archive generee : {odoo_path.resolve()}\n")


# ================================================================
# LANCEMENT PRINCIPAL
# ================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Extracteur d'Images — Mode Unitaire | Mode Excel | Mode Asynchrone",
        formatter_class=argparse.RawTextHelpFormatter
    )
    src = parser.add_mutually_exclusive_group(required=True)
    src.add_argument("--sku",   help="Reference unique, ex: MH03AC-T03RED")
    src.add_argument("--excel", help="Fichier source (doit contenir la colonne 'Internal Reference')")

    parser.add_argument(
        "--urls", nargs="*", metavar="URL",
        help=(
            "Activer le mode de balayage asynchrone [RACE] :\n"
            "  --urls https://site.com/search?q={sku}\n"
            "  --urls   (vide = liste par defaut + Google Images)\n"
            "Utiliser {sku} comme placeholder dans l'URL."
        )
    )
    parser.add_argument(
        "--sites", nargs="+", metavar="URL",
        help=(
            "Ajouter des sites personnalises a la liste de recherche :\n"
            "  --sites https://www.amazon.com/s?k={sku} https://www.ebay.com/sch/i.html?_nkw={sku}\n"
            "Ces URLs seront ajoutees aux URLs par defaut."
        )
    )
    parser.add_argument("--output",   default="./images_dl", help="Repertoire cible")
    parser.add_argument("--headless", action="store_true",   help="Executer en arriere-plan sans interface")
    parser.add_argument("--dry-run",  action="store_true",   help="Effectuer une simulation sans telecharger")

    args        = parser.parse_args()
    use_race    = args.urls is not None or args.sites is not None
    custom_urls = args.urls if args.urls else []
    # Ajouter les sites personnalises
    if args.sites:
        custom_urls.extend(args.sites)
    output_base = Path(args.output)

    if args.excel:
        run_excel_batch(args.excel, output_base, args.headless, args.dry_run, use_race, custom_urls)
    else:
        run_single(args.sku.strip(), output_base, args.headless, args.dry_run, use_race, custom_urls)

    print("  Operation terminee avec succes.\n")


if __name__ == "__main__":
    main()